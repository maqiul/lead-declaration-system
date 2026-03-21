import openpyxl
import json

try:
    wb = openpyxl.load_workbook('d:/lead-declaration-system/backend/src/main/resources/templates/alltemple.xlsx')
    output = []
    output.append("Sheets: " + str(wb.sheetnames))
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        output.append(f"\n--- Sheet: {sheetname} ---")
        for row in range(1, 30):
            for col in range(1, 20):
                val = sheet.cell(row=row, column=col).value
                if val is not None and str(val).strip() != "":
                    output.append(f"R{row}C{col}: {val}")
                
    with open('d:/lead-declaration-system/tmp/layout_all_sheets.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(output))
        
    print("Success")
except Exception as e:
    print(f"Error: {e}")
