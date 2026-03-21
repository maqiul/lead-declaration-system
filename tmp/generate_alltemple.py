import openpyxl

try:
    wb = openpyxl.load_workbook('d:/lead-declaration-system/backend/src/main/resources/templates/alltemple.xlsx')
    sheet = wb.active
    
    def set_val(r, c, v):
        cell = sheet.cell(row=r, column=c)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merge_range in sheet.merged_cells.ranges:
                if cell.coordinate in merge_range:
                    mc_tl = sheet.cell(row=merge_range.min_row, column=merge_range.min_col)
                    mc_tl.value = v
                    return
        else:
            cell.value = v

    set_val(2, 1, "预录入编号：{preEntryNo}")
    set_val(2, 5, "海关编号：{customsNo}")
    set_val(3, 1, "境内发货人 {shipperCompanyCode}")
    set_val(4, 1, "{shipperCompanyName}")
    set_val(4, 5, "{exportCustoms}")
    set_val(4, 7, "{exportDate}")
    set_val(4, 11, "{declarationDate}")
    set_val(4, 14, "{recordNo}")
    set_val(5, 1, "{shipperUniformCode}")
    set_val(7, 1, "{consigneeCompanyName}")
    set_val(7, 5, "{transportMode}")
    set_val(7, 7, "{transportNameAndVoyage}")
    set_val(7, 11, "{billOfLadingNo}")
    set_val(9, 1, "{manufacturer}")
    set_val(9, 5, "{supervisionMode}")
    set_val(9, 7, "{exemptionNature}")
    set_val(9, 11, "{licenseNo}")
    set_val(11, 1, "{contractNo}")
    set_val(11, 5, "{tradeCountry}")
    set_val(11, 7, "{destinationCountry}")
    set_val(11, 11, "{portOfDestination}")
    set_val(11, 14, "{portOfDeparture}")
    set_val(13, 1, "{packageType}")
    set_val(13, 5, "{totalCartons}")
    set_val(13, 6, "{totalGrossWeight}")
    set_val(13, 7, "{totalNetWeight}")
    set_val(13, 9, "{tradeTerm}")
    set_val(13, 11, "{freight}")
    set_val(13, 13, "{premium}")
    set_val(13, 16, "{miscFee}")
    set_val(15, 1, "随附单证1：{attachment1}")
    set_val(15, 8, "随附单证2：{attachment2}")
    set_val(17, 1, "{marksAndRemarks}")
    set_val(19, 1, "{items.no}")
    set_val(19, 2, "{items.hsCode}")
    set_val(19, 4, "{items.name}")
    set_val(19, 7, "{items.quantityStr}")
    set_val(19, 8, "{items.unitPrice}")
    set_val(19, 9, "{items.totalPrice}")
    set_val(19, 10, "{items.currency}")
    set_val(19, 11, "{items.originCountry}")
    set_val(19, 12, "{items.destinationCountry}")
    set_val(19, 14, "{items.sourceRegion}")
    set_val(19, 17, "{items.exemptionType}")
    set_val(20, 4, "{items.specAndModel}")
    set_val(20, 7, "{items.statQuantity}")
    
    wb.save('d:/lead-declaration-system/backend/src/main/resources/templates/alltemple_template.xlsx')
    print("Success: Generated alltemple_template.xlsx")
except Exception as e:
    import traceback
    traceback.print_exc()
