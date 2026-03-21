import openpyxl

try:
    wb = openpyxl.load_workbook('d:/lead-declaration-system/backend/src/main/resources/templates/alltemple.xlsx')
    
    def set_val(sheet, r, c, v):
        cell = sheet.cell(row=r, column=c)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merge_range in sheet.merged_cells.ranges:
                if cell.coordinate in merge_range:
                    mc_tl = sheet.cell(row=merge_range.min_row, column=merge_range.min_col)
                    mc_tl.value = v
                    return
        else:
            cell.value = v

    # --- 1. 报关单 ---
    if '报关单' in wb.sheetnames:
        s1 = wb['报关单']
        set_val(s1, 2, 1, "预录入编号：{preEntryNo}")
        set_val(s1, 2, 5, "海关编号：{customsNo}")
        set_val(s1, 3, 1, "境内发货人 {shipperCompanyCode}")
        set_val(s1, 4, 1, "{shipperCompanyName}")
        set_val(s1, 4, 5, "{exportCustoms}")
        set_val(s1, 4, 7, "{exportDate}")
        set_val(s1, 4, 11, "{declarationDate}")
        set_val(s1, 4, 14, "{recordNo}")
        set_val(s1, 5, 1, "{shipperUniformCode}")
        set_val(s1, 7, 1, "{consigneeCompanyName}")
        set_val(s1, 7, 5, "{transportMode}")
        set_val(s1, 7, 7, "{transportNameAndVoyage}")
        set_val(s1, 7, 11, "{billOfLadingNo}")
        set_val(s1, 9, 1, "{manufacturer}")
        set_val(s1, 9, 5, "{supervisionMode}")
        set_val(s1, 9, 7, "{exemptionNature}")
        set_val(s1, 9, 11, "{licenseNo}")
        set_val(s1, 11, 1, "{contractNo}")
        set_val(s1, 11, 5, "{tradeCountry}")
        set_val(s1, 11, 7, "{destinationCountry}")
        set_val(s1, 11, 11, "{portOfDestination}")
        set_val(s1, 11, 14, "{portOfDeparture}")
        set_val(s1, 13, 1, "{packageType}")
        set_val(s1, 13, 5, "{totalCartons}")
        set_val(s1, 13, 6, "{totalGrossWeight}")
        set_val(s1, 13, 7, "{totalNetWeight}")
        set_val(s1, 13, 9, "{tradeTerm}")
        set_val(s1, 13, 11, "{freight}")
        set_val(s1, 13, 13, "{premium}")
        set_val(s1, 13, 16, "{miscFee}")
        set_val(s1, 15, 1, "随附单证1：{attachment1}")
        set_val(s1, 15, 8, "随附单证2：{attachment2}")
        set_val(s1, 17, 1, "{marksAndRemarks}")
        set_val(s1, 19, 1, "{items.no}")
        set_val(s1, 19, 2, "{items.hsCode}")
        set_val(s1, 19, 4, "{items.name}")
        set_val(s1, 19, 7, "{items.quantityStr}")
        set_val(s1, 19, 8, "{items.unitPrice}")
        set_val(s1, 19, 9, "{items.totalPrice}")
        set_val(s1, 19, 10, "{items.currency}")
        set_val(s1, 19, 11, "{items.originCountry}")
        set_val(s1, 19, 12, "{items.destinationCountry}")
        set_val(s1, 19, 14, "{items.sourceRegion}")
        set_val(s1, 19, 17, "{items.exemptionType}")
        set_val(s1, 20, 4, "{items.specAndModel}")
        set_val(s1, 20, 7, "{items.statQuantity}")

    # --- 2. 申报要素 ---
    if '申报要素' in wb.sheetnames:
        s2 = wb['申报要素']
        set_val(s2, 4, 1, "项号：{items.no}")
        set_val(s2, 5, 1, "商品编码：{items.hsCode}")
        set_val(s2, 6, 1, "商品名称：{items.name}")
        set_val(s2, 7, 1, "0:用途; {items.purpose}")
        set_val(s2, 8, 1, "1:是否录制;{items.isRecorded}")
        set_val(s2, 9, 1, "2:品牌; {items.brand}")
        set_val(s2, 10, 1, "3:型号：{items.model}")
        set_val(s2, 11, 1, "4:品牌类型;{items.brandType}")
        set_val(s2, 12, 1, "5:出口享惠情况;{items.exportPreference}")

    # --- 3. 发票 ---
    if '发票' in wb.sheetnames:
        s3 = wb['发票']
        set_val(s3, 1, 1, "Issuer: {issuerCompanyName}")
        set_val(s3, 2, 1, "{issuerCompanyAddress}")
        set_val(s3, 3, 1, "To: {consigneeCompanyName}")
        set_val(s3, 4, 1, "{consigneeCompanyAddress}")
        set_val(s3, 4, 2, "No:  {invoiceNo}")
        set_val(s3, 4, 5, "Date:  {invoiceDate}")
        set_val(s3, 5, 1, "Transport details             {transportModeEng}")
        set_val(s3, 5, 2, "Terms of payment  {paymentTerms}")
        set_val(s3, 6, 1, "From  {portOfDepartureEng}")
        set_val(s3, 7, 1, "To     {destinationCountryEng}")
        
        set_val(s3, 10, 1, "{items.nameEng}")
        set_val(s3, 10, 2, "{items.quantity}")
        set_val(s3, 10, 3, "{items.unitEng}")
        set_val(s3, 10, 4, "{items.currencyEng}")
        set_val(s3, 10, 5, "{items.unitPrice}")
        set_val(s3, 10, 6, "{items.currencyEng}")
        
        set_val(s3, 18, 1, "SAY {totalAmountEng} ONLY")
        set_val(s3, 19, 1, "TOTAL PACKED IN {totalCartons} {packageTypeEng}")

    # --- 4. 装箱单 ---
    if '装箱单' in wb.sheetnames:
        s4 = wb['装箱单']
        set_val(s4, 1, 1, "Issuer: {issuerCompanyName}")
        set_val(s4, 2, 1, "{issuerCompanyAddress}")
        set_val(s4, 3, 1, "To: {consigneeCompanyName}")
        set_val(s4, 4, 1, "{consigneeCompanyAddress}")
        set_val(s4, 4, 2, "No:  {packingListNo}")
        set_val(s4, 4, 6, "Date:  {packingListDate}") # Note: col 6 instead of 5
        set_val(s4, 5, 1, "Transport details             {transportModeEng}")
        set_val(s4, 5, 2, "Terms of payment  {paymentTerms}")
        set_val(s4, 6, 1, "From  {portOfDepartureEng}")
        set_val(s4, 7, 1, "To      {destinationCountryEng}")
        
        set_val(s4, 9, 1, "{items.nameEng}")
        set_val(s4, 9, 2, "{items.quantity}")
        set_val(s4, 9, 3, "{items.unitEng}")
        set_val(s4, 9, 4, "{items.cartons}")
        set_val(s4, 9, 5, "{packageTypeEng}")
        set_val(s4, 9, 6, "{items.grossWeight}")
        set_val(s4, 9, 8, "{items.netWeight}")
        set_val(s4, 9, 10, "{items.volume}")
        
        set_val(s4, 18, 1, "TOTAL PACKED IN {totalCartons} {packageTypeEng}")

    wb.save('d:/lead-declaration-system/backend/src/main/resources/templates/alltemple_template.xlsx')
    print("Success: Generated alltemple_template.xlsx for all 4 sheets")
except Exception as e:
    import traceback
    traceback.print_exc()
